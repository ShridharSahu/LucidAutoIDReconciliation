from time import sleep
import sys
from pathlib import Path
import pyinputplus as pyip
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import ctypes

# Display Always on
ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)

# Initial Message
print('-------------- Auto Lucid Reconciliation --------------\n')
print('This is a web scrapper tool that automates the process for Ids reconciliation in Lucid')
print('This is created by Shridhar Sahu. Please contact him in case of any queries\n')
sleep(1)
print('Prerequisite:')
print('1. Google Chrome')
print('2. Lucid Project ID and Project Name in InputSheet.xlsx')
print('3. Final Ids in the IdsToBePasted.txt file')
print('4. All input files should be closed')
print('5. No input files should be renamed\n')
sleep(1)

# Setting Excel
InputFile = Path.cwd()/'InputSheet.xlsx'
workbook = openpyxl.load_workbook(InputFile)
sheet = workbook['Sheet1']
try:
    workbook.save(InputFile)
except:
    print('Please close the InputSheet.xlsx file')
    sleep(20)
    sys.exit()

# Input for email ID, Password and Delay
print('User Input')
EmailID = pyip.inputEmail('Email ID (Group email ID only): ')
Password = pyip.inputPassword(prompt='Password: ', mask='*')
Delay = pyip.inputInt('Enter delay in seconds (10-100): ', min=10, max=100)

# Setting Chrome up
chrome_options = webdriver.ChromeOptions()
# chrome_options.headless = True
chrome_options.add_experimental_option('prefs', {'download.default_directory': str(Path.cwd()) + '\DataExport'})
driver = webdriver.Chrome(executable_path=Path.cwd() / 'Resources/chromedriver.exe', options=chrome_options)

# Reading Project details from excel
for r in range(2, sheet.max_row +1):
    if sheet.cell(row=r, column=1).value in [None, '']:
        break
    ProjectID = sheet.cell(row=r, column=1).value
    ProjectName = (str(sheet.cell(row=r, column=2).value)[0:30]).strip()

    # Closing program if no more project Id is present
    if str(ProjectID).strip() == '':
        driver.quit()
        workbook.save(InputFile)
        print('\nIds Reconciliation is done. Please check status for all setups in InputSheet.xlsx')
        sleep(20)
        sys.exit()

    try:

        # Setting wait time and opening url
        driver.maximize_window()
        driver.implicitly_wait(Delay)
        driver.get('https://login.samplicio.us')

        # Log In Page
        if driver.title == 'Log In - Lucid':
            driver.find_element(By.NAME, 'email').send_keys(EmailID)
            driver.find_element(By.NAME, 'password').send_keys(Password)
            driver.find_element(By.NAME, 'submit').click()

        # Reset all search, Click on All (wait till you get a different all text), Search PID and click the PName
        driver.find_element(By.XPATH, '//*[@id="root"]/div[2]/main/div/div[2]/button').click()
        elem_all = driver.find_element(By.ID, 'tab-all')
        previousElemAllText, currentElemAllText = elem_all.text, elem_all.text
        inputDelay = 0
        while (inputDelay <= Delay) and (previousElemAllText == currentElemAllText):
            sleep(1)
            inputDelay += 1
            previousElemAllText, currentElemAllText = currentElemAllText, elem_all.text
        elem_all.click()
        driver.find_element(By.XPATH, '//*[@id="root"]/div[2]/main/div/form/div[1]/label/div/input').send_keys(ProjectID)
        driver.find_element(By.XPATH, '//*[@id="root"]/div[2]/main/div/form/div[1]/label/div/span/button').click()
        driver.find_element(By.PARTIAL_LINK_TEXT, ProjectName).click()

        # Click the Reconciliation Tab, Click on Data Analysis, Update the field start date and download the Data Analysis Report
        driver.find_element(By.LINK_TEXT, 'Reconciliations').click()
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_btnDataAnalysis').click()
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_tbFromDate').clear()
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_tbFromDate').send_keys('01/01/2019')
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnGenerate').click()

        # Click on Reconciliation, Upload file and click finish multiple times
        # Check if there are pop-ups for bad id. In case final button is visible there is no popup
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_btnReconciliations').click()
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_fuUploads').send_keys(str(Path.cwd()) + '\IdsToBePasted.txt')
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep1Continue').click()
        driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep2PreprocessContinue').click()
        elem_Step2Finish = driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep2Finish')
        if elem_Step2Finish.is_displayed():
            elem_Step2Finish.click()
        else:
            driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep2TerminateYes').click()
            elem_Step2Finish = driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep2Finish')
            if elem_Step2Finish.is_displayed():
                elem_Step2Finish.click()
            else:
                driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnConfirmStep2TerminateYes').click()
                driver.find_element(By.ID, 'ctl00_ctl00_ContentPlaceHolder1_ContentItems_btnStep2Finish').click()

        sheet.cell(row=r, column=3).value = 'Pass'
        workbook.save(InputFile)

    except:

        sheet.cell(row=r, column=3).value = 'Fail'
        workbook.save(InputFile)

driver.quit()
workbook.save(InputFile)
print('\nIds Reconciliation is done. Please check status for all setups in InputSheet.xlsx')
sleep(20)
ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)

